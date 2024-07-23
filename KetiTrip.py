# KetiTrip.py
import os
import sys
import re
import openpyxl
from typing import List
from os import listdir
from os.path import isfile, join
from datetime import date, datetime, timedelta

WorkingFolder = "C:/Users/ParkPusik/OneDrive - OPENLAB/MOBILITY/_참여율&연구수당,외부수당/국외출장"

FilePrefix = "출장관리_출장현황(출장자)_"

# Define the team members
team_a = ["신대교", "윤상훈", "장수현", "장성현", "임기택", "안병만", "장준혁"]
team_b = ["민경원", "손행선", "이선영", "박진만", "심영보", "조용현", "성기호", "유재웅"]
team_c = ["정한균", "진성근"]
team_d = ["박부식"]

class Team:

    def __init__(self, name):
        self.name = name
        self.total_count = 0
        self.total_cost = 0


class TripPerPerson:

    def __init__(self, path_name):
        self.path_name = path_name
        # Second string of path_name
        self.name = path_name.split('_')[2].split('.')[0]

        # self.name 이 team_a에 있는지 확인
        if self.name in team_a:
            self.team = "모쏠"
        elif self.name in team_b:
            self.team = "자율"
        elif self.name in team_c:
            self.team = "통신"
        elif self.name in team_d:
            self.team = "부식"
        else:
            self.team = "Unknown"
            
        self.year_start = 9999
        self.year_end = 0
        self.cost = 0
        self.count = 0

        # Define variable to load the dataframe
        self.wb = openpyxl.load_workbook(path_name)

        # Define variable to read sheet
        self.ws = self.wb.active

        offset = 1
        # Iterate the loop to read the cell values
        for row_idx in range(offset, self.ws.max_row):
            # print("row_idx:", row_idx, "self.ws.cell(row=row_idx, column=2).value:", self.ws.cell(row=row_idx, column=2).value)
            # print("row_idx:", row_idx, "self.ws.cell(row=row_idx, column=15).value:", self.ws.cell(row=row_idx, column=15).value)
            
            # Second column value is 272 or not, 국외여비
            str = self.ws.cell(row=row_idx, column=1).value
            # 앞에 4자리를 추출해 숫자로 저장
            if str[:4].isdigit():
                year = int(str[:4])
                # 최소, 최대 연도를 저장
                if year < self.year_start:
                    self.year_start = year
                if year > self.year_end:
                    self.year_end = year
            else:
                continue

            cost = int(self.ws.cell(row=row_idx, column=16).value)
            self.cost += cost
            self.count += 1

        self.wb.close()


trips: List[TripPerPerson] = []
TeamA = Team("모쏠")
TeamB = Team("자율")
TeamC = Team("통신")
TeamD = Team("부식")


def func_init():
    global WorkingFolder

    os.chdir(WorkingFolder)
    print("os.getcwd():", os.getcwd())


def func_get_info_from_files():
    # List all files in the folder
    Files = [f for f in listdir(WorkingFolder) if isfile(join(WorkingFolder, f))]

    for f in Files:
        if f.startswith(FilePrefix):
            new_trip = TripPerPerson(f)

            trips.append(new_trip)
            
def func_sum_trips():
    for t in trips:

        if t.team == TeamA.name:
            TeamA.total_cost += t.cost
            TeamA.total_count += t.count
        elif t.team == TeamB.name:
            TeamB.total_cost += t.cost
            TeamB.total_count += t.count
        elif t.team == TeamC.name:
            TeamC.total_cost += t.cost
            TeamC.total_count += t.count
        elif t.team == TeamD.name:
            TeamD.total_cost += t.cost
            TeamD.total_count += t.count
        else:
            print("Unknown team:", t.team)

def print_trips():
    for t in trips:
        print("이름:", t.name, ", 횟수:", f"{t.count:2}", ", 비용:", f"{t.cost:12,}")
    print("---")


def print_teams():
    print("팀명     횟수       비용")
    print(TeamA.name, f"{TeamA.total_count:3,}", f"{TeamA.total_cost:12,}")
    print(TeamB.name, f"{TeamB.total_count:3,}", f"{TeamB.total_cost:12,}")
    print(TeamC.name, f"{TeamC.total_count:3,}", f"{TeamC.total_cost:12,}")
    print(TeamD.name, f"{TeamD.total_count:3,}", f"{TeamD.total_cost:12,}")
    print("---")


def main():
    print("Hello, KETI Oversee Trip Analysis")

    func_init()
    
    func_get_info_from_files()

    func_sum_trips()
    
    print_trips()

    print_teams()

    
main()
