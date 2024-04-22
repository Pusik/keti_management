import re
import openpyxl
from typing import List
from os import listdir
from os.path import isfile, join
from datetime import date, timedelta

# 과제 예실대비표 클래스
class ProjectBudget:
    def __init__(self, target_year, path_name):
        self.target_year = target_year
        self.path_name = path_name

        self.project_file_id = ""    # 계정번호(파일명)
        self.project_cur_id = ""     # 계정번호
        self.project_base_id = ""    # 과제번호

        self.expense_internal = 0   # 내부인건비
        self.expense_external = 0   # 계약직내부인건비
        self.expense_overhead = 0   # 간접비
        self.expense_internal_target = 0   # 내부인건비
        self.expense_external_target = 0   # 계약직내부인건비
        self.expense_overhead_target = 0   # 간접비
        self.date_begin = date(2000, 1, 1)     # 과제시작일
        self.date_end = date(2000, 1, 1)       # 과제종료일
        self.months = 0
        self.months_target = 0

        #print(path_name)

        # Find a string after startswith
        self.id = self.project_file_id = re.findall(r'\((.*?)\)', path_name)[0]

        # Define variable to load the dataframe
        self.wb = openpyxl.load_workbook(path_name)

        # Define variable to read sheet
        self.ws = self.wb.active

        offset = 1
        # Iterate the loop to read the cell values
        for row_idx in range(offset, self.ws.max_row):
            if self.ws.cell(row=row_idx, column=2).value == "121":
                self.expense_internal = int(self.ws.cell(row=row_idx, column=7).value)
                #print("내부인건비", self.ws.cell(row=row_idx, column=7).value)
            elif self.ws.cell(row=row_idx, column=2).value == "201":
                self.expense_external = int(self.ws.cell(row=row_idx, column=7).value)
                #print("계약직내부인건비", self.ws.cell(row=row_idx, column=7).value)
            elif self.ws.cell(row=row_idx, column=2).value == "123":
                self.expense_overhead = int(self.ws.cell(row=row_idx, column=7).value)
                #print("간접비", self.ws.cell(row=row_idx, column=7).value)

        self.wb.close()

    def show(self):
        print(self.path_name)
        print("과제번호(파일):", self.project_file_id)
        print("과제번호:", self.project_base_id)
        print("계정번호:", self.project_cur_id)
        print("과제시작일:", self.date_begin)
        print("과제종료일:", self.date_end)
        print("월수/월수({0}):".format(self.target_year), self.months, self.months_target)
        print("내부인건비:", "{:,}".format(self.expense_internal))
        print("계약직내부인건비:", "{:,}".format(self.expense_external))
        print("간접비:", "{:,}".format(self.expense_overhead))
        print("내부인건비({0}):".format(self.target_year), "{:,}".format(self.expense_internal_target))
        print("계약직내부인건비({0}):".format(self.target_year), "{:,}".format(self.expense_external_target))
        print("간접비({0}):".format(self.target_year), "{:,}".format(self.expense_overhead_target))

# 위촉연구원 발령정보 클래스
class Contract:
    def __init__(self, target_year, id, name, project_id, date_start, date_end, working_time, expense_per_hour, expense, expense_holiday, expense_extra):
        self.target_year = target_year  # 당해년도
        self.id = id    # 사번
        self.name = name    # 성명
        self.project_id = project_id    # 계정번호
        self.date_start = date_start    # 발령 시작일
        self.date_end = date_end    # 발령 종료일
        
        if self.date_start > self.date_end:
            print("Error: Date Start(", self.date_start, ") is greater than Date End(", self.date_end, ")")
            return 
        # 발령 기간[월]
        diff = self.date_end - self.date_start
        self.months = int(diff.days / 30)
        # 당해년도 근무 월수
        if self.date_start.year > self.target_year:
            print("Error: Date Start(", self.date_start, ") is greater than Target Year(", self.target_year, ")")
            return
        if self.date_start.year < self.target_year:
            date1 = date(self.target_year, 1, 1)
        else:
            date1 = self.date_start
        if self.date_end.year < self.target_year:
            print("Error: Date End(", self.date_end, ") is less than than Target Year(", self.target_year, ")")
            return
        if self.date_end.year > self.target_year:
            date2 = date(self.target_year, 12, 31)
        else:
            date2 = self.date_end
        diff = date2 - date1
        self.target_months = int(diff.days / 30)

        self.working_time = int(working_time)   # 발령 계약시간
        self.expense_per_hour = int(expense_per_hour) # 시간당 단가
        if expense_holiday == None:    # 주유수당
            self.monthly_expense_holiday = 0
        else:
            self.monthly_expense_holiday = int(expense_holiday)
        self.monthly_expense = int(expense)   # 월정액 = (시간당 단가) x (발령 계약시간) + (주유수당)
        if expense_extra == None:  # 추가부담
            self.montly_expense_extra = 0
        else:
            self.monthly_expense_extra = int(expense_extra)

        # 당해년도 
        self.target_total_salary = self.monthly_expense * self.target_month
        self.target_total_expense = (self.monthly_expense + self.monthly_expense_extra) * self.target_month

    def show(self):
        print("성명:", self.name, "사번:", self.id, "계정번호:", self.project_id)
        print("과제시작:", self.date_start, "과제종료:", self.date_end, "발령월수:", self.months)
        print("근무시간:", self.working_time, "단가:", "{:,}".format(self.expense_per_hour), "주유수당:", "{:,}".format(self.expense_holiday), "월정액:", "{:,}".format(self.expense_monthly), "추가비용:", "{:,}".format(self.expense_extra))
        print("당해년도 근무월수:", self.target_months, "당해년도 월정액 합:", "{:,}".format(self.target_total_salary), "당해년도 지출 합:", "{:,}".format(self.target_total_expense))
        print("-------------------------------------")

# 위촉연구원 클래스
class Employee:
    def __init__(self, target_year, id, name):
        self.target_year = target_year
        self.id = id       # 사번
        self.name = name    # 성명

        self.target_monthly_salary = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0] # 월별 연봉
        self.target_monthly_expense = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0] # 월별 지출

        self.target_total_salary = 0  # 연봉
        self.target_total_expense = 0  # 지출

        self.contracts: List[Contract] = []  # 계약
        
    def add(self, contract: Contract):
        self.contracts.append(contract)

        for i in range(1, 12):
            start_month = (self.target_year - contract.date_start.year) * 12 + contract.date_start.month
            end_month = (self.target_year - contract.date_end.year) * 12 + contract.date_end.month

            if start_month <= i and i <= end_month:
                self.target_monthly_salary[i-1] += contract.target_total_salary
                self.target_monthly_expense[i-1] += contract.target_total_expense

                self.target_total_salary += contract.target_total_salary
                self.target_total_expense += contract.target_total_expense

    def show(self):
        print("성명:", self.name, "사번:", self.id, "급여:", "{:,}".format(self.target_total_salary), "총지출:", "{:,}".format(self.target_total_expense))
        for i in range(1, 12):
            print("{0}월 연봉:{1:,} 지출:{2:,}".format(i, self.target_monthly_salary[i-1], self.target_monthly_expense[i-1]))
        print("-------------------------------------")

