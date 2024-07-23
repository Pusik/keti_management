# KetiTrip.py
import os
import sys
import re
import openpyxl
from typing import List
from os import listdir
from os.path import isfile, join
from datetime import date, datetime, timedelta

#WorkingFolder = "C:/Users/ParkPusik/OneDrive - OPENLAB/MOBILITY/_참여율&연구수당,외부수당/예실대비표20231231"
WorkingFolder = "C:/Users/ParkPusik/OneDrive - OPENLAB/MOBILITY/_참여율&연구수당,외부수당/예실대비표20240723"

FilePrefix = "예실대비표 "

# Define the team members
team_a = ["신대교", "윤상훈", "장수현", "장성현", "임기택", "안병만", "장준혁"]
team_b = ["민경원", "손행선", "이선영", "박진만", "심영보", "조용현", "성기호", "유재웅"]
team_c = ["정한균", "진성근"]
team_d = ["박부식"]

class Team:

    def __init__(self, name):
        self.name = name
        self.total_cost_oversee_trip = 0
        self.total_cost_meeting = 0
        self.total_salary_regular = 0
        self.total_salary_irregular = 0
        self.total_income_indirect = 0

class Employee:

    def __init__(self, name):
        self.name = name
        # self.name 이 team_a에 있는지 확인
        if self.name in team_a:
            self.team = "모쏠"
        elif self.name in team_b:
            self.team = "자율"
        elif self.name in team_c:
            self.team = "통신"
        elif self.name in team_d:
            self.team = "부식"
        else:
            self.team = "Unknown"

        self.total_cost_oversee_trip = 0
        self.total_cost_meeting = 0
        self.total_salary_regular = 0
        self.total_salary_irregular = 0
        self.total_income_indirect = 0

class ProjectBudget:

    def __init__(self, path_name):
        self.path_name = path_name
        # Second string of path_name
        self.name = path_name.split(' ')[1]

        # Third 8-digit string of path_name
        self.code = path_name.split(' ')[2].split('.')[0]

        self.cost_oversee_trip = -1
        self.cost_meeting = -1
        self.salary_regular = -1
        self.salary_irregular = -1
        self.income_indirect = -1

        # Define variable to load the dataframe
        self.wb = openpyxl.load_workbook(path_name)

        # Define variable to read sheet
        self.ws = self.wb.active

        offset = 1
        # Iterate the loop to read the cell values
        for row_idx in range(offset, self.ws.max_row):
            # print("row_idx:", row_idx, "self.ws.cell(row=row_idx, column=2).value:", self.ws.cell(row=row_idx, column=2).value)
            # print("row_idx:", row_idx, "self.ws.cell(row=row_idx, column=15).value:", self.ws.cell(row=row_idx, column=15).value)
            
            # Second column value is 272 or not, 국외여비
            if self.ws.cell(row=row_idx, column=2).value == "272":
                string_value = ""

                if self.cost_oversee_trip != -1:
                    print("Error: Multiple Oversea Trip Costs")
                    break
                # Column O 's value is the oversea trip cost
                string_value = self.ws.cell(row=row_idx, column=15).value
                # convert string to integer
                self.cost_oversee_trip = int(string_value)
                # print("self.cost_oversee_trip:", self.cost_oversee_trip)

            # Second column value is 410 or not, 회의비
            if self.ws.cell(row=row_idx, column=2).value == "410":
                if self.cost_meeting != -1:
                    print("Error: Multiple Meeting Costs")
                    break
                # Column O 's value is the oversea trip cost
                string_value = self.ws.cell(row=row_idx, column=15).value
                self.cost_meeting = int(string_value)
                # print("self.cost_meeting:", self.cost_meeting)

            # Second column value is 121 or not, 내부인건비
            if self.ws.cell(row=row_idx, column=2).value == "121":
                if self.salary_regular != -1:
                    print("Error: Multiple Salary Regular")
                    break
                # Column O 's value is the oversea trip cost
                string_value = self.ws.cell(row=row_idx, column=15).value
                self.salary_regular = int(string_value)
                # print("self.salary_regular:", self.salary_regular)
            
            # Second column value is 201 or not, 외부인건비
            if self.ws.cell(row=row_idx, column=2).value == "201":
                if self.salary_irregular != -1:
                    print("Error: Multiple Salary Irregular")
                    break
                # Column O 's value is the oversea trip cost
                string_value = self.ws.cell(row=row_idx, column=15).value
                self.salary_irregular = int(string_value)
                # print("self.salary_irregular:", self.salary_irregular)
            
            # Second column value is 123 or not, 간접비
            if self.ws.cell(row=row_idx, column=2).value == "123":
                if self.income_indirect != -1:
                    print("Error: Multiple Income Indirect")
                    break
                # Column O 's value is the oversea trip cost
                string_value = self.ws.cell(row=row_idx, column=15).value
                self.income_indirect = int(string_value)
                # print("self.income_indirect:", self.income_indirect)
        
        if self.cost_oversee_trip == -1:
            self.cost_oversee_trip = 0
        if self.cost_meeting == -1:
            self.cost_meeting = 0
        if self.salary_regular == -1:  
            self.salary_regular = 0
        if self.salary_irregular == -1:
            self.salary_irregular = 0
        if self.income_indirect == -1:
            self.income_indirect = 0

        self.wb.close()


budgets: List[ProjectBudget] = []
employyes: List[Employee] = []
TeamA = Team("모쏠")
TeamB = Team("자율")
TeamC = Team("통신")
TeamD = Team("부식")

def func_init():
    global WorkingFolder

    os.chdir(WorkingFolder)
    print("os.getcwd():", os.getcwd())


def func_get_info_from_files():
    # List all files in the folder
    Files = [f for f in listdir(WorkingFolder) if isfile(join(WorkingFolder, f))]

    for f in Files:
        if f.startswith(FilePrefix):
            new_budget = ProjectBudget(f)

            budgets.append(new_budget)
            
def func_sum_budgets():
    for b in budgets:
        # Search the name in the employee list
        found = False
        for e in employyes:
            if b.name == e.name:
                found = True
                break

        if not found:
            new_employee = Employee(b.name)
            new_employee.total_cost_oversee_trip = b.cost_oversee_trip
            new_employee.total_cost_meeting = b.cost_meeting
            new_employee.total_salary_regular = b.salary_regular
            new_employee.total_salary_irregular = b.salary_irregular
            new_employee.total_income_indirect = b.income_indirect
            
            employyes.append(new_employee)
        else:  
            e.total_cost_oversee_trip += b.cost_oversee_trip
            e.total_cost_meeting += b.cost_meeting
            e.total_salary_regular += b.salary_regular
            e.total_salary_irregular += b.salary_irregular
            e.total_income_indirect += b.income_indirect

    # Team별로 합산
    for e in employyes:
        if e.team == "모쏠":
            TeamA.total_cost_oversee_trip += e.total_cost_oversee_trip
            TeamA.total_cost_meeting += e.total_cost_meeting
            TeamA.total_salary_regular += e.total_salary_regular
            TeamA.total_salary_irregular += e.total_salary_irregular
            TeamA.total_income_indirect += e.total_income_indirect
        elif e.team == "자율":
            TeamB.total_cost_oversee_trip += e.total_cost_oversee_trip
            TeamB.total_cost_meeting += e.total_cost_meeting
            TeamB.total_salary_regular += e.total_salary_regular
            TeamB.total_salary_irregular += e.total_salary_irregular
            TeamB.total_income_indirect += e.total_income_indirect
        elif e.team == "통신":
            TeamC.total_cost_oversee_trip += e.total_cost_oversee_trip
            TeamC.total_cost_meeting += e.total_cost_meeting
            TeamC.total_salary_regular += e.total_salary_regular
            TeamC.total_salary_irregular += e.total_salary_irregular
            TeamC.total_income_indirect += e.total_income_indirect
        elif e.team == "부식":
            TeamD.total_cost_oversee_trip += e.total_cost_oversee_trip
            TeamD.total_cost_meeting += e.total_cost_meeting
            TeamD.total_salary_regular += e.total_salary_regular
            TeamD.total_salary_irregular += e.total_salary_irregular
            TeamD.total_income_indirect += e.total_income_indirect


def print_budgets():
    for b in budgets:
        print(b.name, b.code, b.cost_oversee_trip, b.cost_meeting)

def print_employees():
    for e in employyes:
        print(e.name, f"{e.total_cost_oversee_trip:15,}", f"{e.total_cost_meeting:15,}")
    print("---")

def print_teams():
    # print static string with a fixed length

    print("팀명     국외여비       회의비     내부인건비   계약직인건비         간접비")
    print(TeamA.name, f"{TeamA.total_cost_oversee_trip:12,}", f"{TeamA.total_cost_meeting:12,}", f"{TeamA.total_salary_regular:14,}", f"{TeamA.total_salary_irregular:14,}", f"{TeamA.total_income_indirect:14,}")
    print(TeamB.name, f"{TeamB.total_cost_oversee_trip:12,}", f"{TeamB.total_cost_meeting:12,}", f"{TeamB.total_salary_regular:14,}", f"{TeamB.total_salary_irregular:14,}", f"{TeamB.total_income_indirect:14,}")
    print(TeamC.name, f"{TeamC.total_cost_oversee_trip:12,}", f"{TeamC.total_cost_meeting:12,}", f"{TeamC.total_salary_regular:14,}", f"{TeamC.total_salary_irregular:14,}", f"{TeamC.total_income_indirect:14,}")
    print(TeamD.name, f"{TeamD.total_cost_oversee_trip:12,}", f"{TeamD.total_cost_meeting:12,}", f"{TeamD.total_salary_regular:14,}", f"{TeamD.total_salary_irregular:14,}", f"{TeamD.total_income_indirect:14,}")
    print("---")


def main():
    print("Hello, KETI Budget Analysis")

    func_init()
    
    func_get_info_from_files()

    func_sum_budgets()

    # print_budgets()
    print_employees()

    print_teams()

    # sum of total salary regular
    total_salary_regular = TeamA.total_salary_regular + TeamB.total_salary_regular + TeamC.total_salary_regular + TeamD.total_salary_regular
    total_income_indirect = TeamA.total_income_indirect + TeamB.total_income_indirect + TeamC.total_income_indirect + TeamD.total_income_indirect
    total_income = total_salary_regular + total_income_indirect
    print("인건비흡수:", f"{total_salary_regular:,}", "원, 간접비:", f"{total_income_indirect:,}", "원, 총수익:", f"{total_income:,}")
    print("")
    
main()
