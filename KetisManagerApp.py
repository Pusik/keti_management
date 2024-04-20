import os
import sys
import re
import openpyxl
from typing import List
from os import listdir
from os.path import isfile, join
from datetime import date, timedelta
from KetiManager import *

TargetYear = 2024
TargetYearBegin = date(TargetYear, 1, 1)
TargetYearEnd = date(TargetYear, 12, 31)

WorkingFolder = ""

FileBudget = "연구_예산관리_예실대비표(4P4C3369)_Button(엑셀).xlsx"
FileIrregularEmployee = "인사_비상근전문계약직_과제참여현황_Button(엑셀다운).xlsx"
WholeProjectInfo = "연구_과제현황_과제종합정보_Button(엑셀데이터다운).xlsx"

Projects: List[ProjectBudget] = []  # 과제 예실대비표 정보
Employees: List[Employee] = []  # 위촉연구원 정보

def func_init():
    global WorkingFolder
    WorkingFolder = "." if len(sys.argv) <= 1 else sys.argv[1]
    #WorkingFolder = "C:/Users/ParkPusik/OneDrive - OPENLAB/MOBILITY/____센터장App/20231231"
    print("WorkingFolder = '", WorkingFolder, "'")

    global TargetYear
    global TargetYearBegin
    global TargetYearEnd

    global FileIrregularEmployee
    global WholeProjectInfo

    global Projects
    global Employees

def func_project_budget():
    # List all files in the folder

    
    Files = [f for f in listdir(WorkingFolder) if isfile(join(WorkingFolder, f))]

    for f in Files:
        if f.startswith("연구_예산관리_예실대비표("):
            ##print(f)
            
            proj = ProjectBudget(TargetYear, WorkingFolder + "/" + f)
            
            Projects.append(proj)

    # Define the beginning/ending date of the project
    wb = openpyxl.load_workbook(WorkingFolder + "/" + WholeProjectInfo)
    ws = wb.active

    offset = 2
    for proj in Projects:
        for row_idx in range(offset, ws.max_row):
            '''
            print("1", ws.cell(row=row_idx, column=1).value)
            print("2", ws.cell(row=row_idx, column=2).value)
            print("3", ws.cell(row=row_idx, column=3).value)
            print("4", ws.cell(row=row_idx, column=4).value)
            print("5", ws.cell(row=row_idx, column=5).value)
            print("6", ws.cell(row=row_idx, column=6).value)
            print("7", ws.cell(row=row_idx, column=7).value)
            print("8", ws.cell(row=row_idx, column=8).value)
            '''
            #print(ws.cell(row=row_idx, column=9).value, proj.project_cur_id)

            if ws.cell(row=row_idx, column=9).value == proj.project_file_id:
                proj.project_cur_id = proj.project_file_id
                #print("과제시작일", ws.cell(row=row_idx, column=11).value)
                #print("과제종료일", ws.cell(row=row_idx, column=12).value)
                str_date = ws.cell(row=row_idx, column=11).value
                str_date = str_date.replace(".", "")
                proj.date_begin = date(int(str_date[0:4]), int(str_date[4:6]), int(str_date[6:8]))

                str_date = ws.cell(row=row_idx, column=12).value
                str_date = str_date.replace(".", "")
                proj.date_end = date(int(str_date[0:4]), int(str_date[4:6]), int(str_date[6:8]))

                # 과제번호
                proj.project_base_id = ws.cell(row=row_idx, column=1).value
                break
        # No match
        #print("No match for ", proj.project_cur_id)

        for row_idx in range(offset, ws.max_row):
            #print(ws.cell(row=row_idx, column=1).value, proj.project_cur_id)

            if ws.cell(row=row_idx, column=1).value == proj.project_file_id:
                proj.project_cur_id = ""
                #print("과제시작일", ws.cell(row=row_idx, column=11).value)
                #print("과제종료일", ws.cell(row=row_idx, column=12).value)
                str_date = ws.cell(row=row_idx, column=11).value
                str_date = str_date.replace(".", "")
                proj.date_begin = date(int(str_date[0:4]), int(str_date[4:6]), int(str_date[6:8]))

                str_date = ws.cell(row=row_idx, column=12).value
                str_date = str_date.replace(".", "")
                proj.date_end = date(int(str_date[0:4]), int(str_date[4:6]), int(str_date[6:8]))

                # 과제번호
                proj.project_base_id = ws.cell(row=row_idx, column=1).value
                break

    # Remove the project with 0 months
    for proj in Projects:
        # Calculate the month from the date_begin to date_end
        diff = proj.date_end - proj.date_begin
        months = int(diff.days / 30)
        if months == 0:
            Projects.remove(proj)
        else:
            proj.months = months

    # Remove the project out of the target year
    for proj in Projects:
        if proj.date_end < TargetYearBegin or proj.date_begin > TargetYearEnd:
            #print("Out of range", proj.project_base_id, proj.date_begin, proj.date_end)
            Projects.remove(proj)

    # Calculate the monthly expenses
    for proj in Projects:
        monthly_expenses_internal = int(proj.expenses_internal / proj.months)
        monthly_expenses_external = int(proj.expenses_external / proj.months)
        monthly_expenses_overhead = int(proj.expenses_overhead / proj.months)

        if proj.date_begin <= TargetYearBegin and proj.date_end <= TargetYearEnd:
            diff = proj.date_end - TargetYearBegin
            proj.months_target = int(diff.days / 30)
        elif proj.date_begin <= TargetYearBegin and proj.date_end >= TargetYearEnd:  
            proj.months_target = 12
        elif proj.date_begin >= TargetYearBegin and proj.date_end >= TargetYearEnd:
            diff = TargetYearEnd - proj.date_begin
            proj.months_target = int(diff.days / 30)   
        elif proj.date_begin >= TargetYearBegin and proj.date_end <= TargetYearEnd:
            diff = proj.date_end - proj.date_begin
            proj.months_target = int(diff.days / 30)
        else:
            print("Error", proj.project_cur_id, proj.date_begin, proj.date_end)

        proj.expenses_internal_target = monthly_expenses_internal * proj.months_target
        proj.expenses_external_target = monthly_expenses_external * proj.months_target
        proj.expenses_overhead_target = monthly_expenses_overhead * proj.months_target


    #for proj in Projects:
    #    proj.show()

    # Duplication check
    #dup = 0
    #for proj1 in Projects:
    #    for proj2 in Projects:
    #        if proj1.project_cur_id == proj2.project_file_id:
    #            dup += 1
    #            break
    #print("Total", len(Projects), "Duplication: ", dup)

    # Summary of projects
    total_expenses_internal_target = 0
    total_expenses_external_target = 0
    total_expenses_overhead_target = 0

    for proj in Projects:
        # Acculmulate the expenses and overhead
        total_expenses_internal_target += proj.expenses_internal_target
        total_expenses_external_target += proj.expenses_external_target
        total_expenses_overhead_target += proj.expenses_overhead_target

    # Format comma separated number
    print("Total 내부인건비: ", "{:,}".format(total_expenses_internal_target))
    print("Total 계약직내부인건비: ", "{:,}".format(total_expenses_external_target))
    print("Total 간접비: ", "{:,}".format(total_expenses_overhead_target))
    print("Total OH", "{:,}".format(total_expenses_internal_target + total_expenses_overhead_target))


def func_employee():
    # Write Projects into a excel file
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(["과제번호(파일)", "과제번호", "계정번호", "과제시작일", "과제종료일", "월수", "월수({TargetYear})", "내부인건비", "계약직내부인건비", "간접비", "내부인건비{0}".format(TargetYear), "계약직내부인건비{0}".format(TargetYear), "간접비{0}".format(TargetYear)])
    for proj in Projects:
        ws.append([proj.project_file_id, proj.project_base_id, proj.project_cur_id, proj.date_begin, proj.date_end, proj.months, proj.months_target, proj.expenses_internal, proj.expenses_external, proj.expenses_overhead, proj.expenses_internal_target, proj.expenses_external_target, proj.expenses_overhead_target])   

    wb.save(WorkingFolder + "/Summary_Projects" + str(TargetYear) + ".xlsx")
    print("Summary_Projects" + str(TargetYear) + ".xlsx is saved")

    wb.close()


    # Write Projects into a excel file
    wb = openpyxl.load_workbook(WorkingFolder + "/" + FileIrregularEmployee)
    ws = wb.active

    offset = 3
    for row_idx in range(offset, ws.max_row):
        '''
        print("1", ws4.cell(row=row_idx+offset, column=1).value)
        print("2", ws4.cell(row=row_idx+offset, column=2).value)
        print("3", ws4.cell(row=row_idx+offset, column=3).value)
        print("4", ws4.cell(row=row_idx+offset, column=4).value)
        print("5", ws4.cell(row=row_idx+offset, column=5).value)
        print("6", ws4.cell(row=row_idx+offset, column=6).value)
        print("7", ws4.cell(row=row_idx+offset, column=7).value)
        print("8", ws4.cell(row=row_idx+offset, column=8).value)
        print("9", ws4.cell(row=row_idx+offset, column=9).value)
        print("10", ws4.cell(row=row_idx+offset, column=10).value)
        print("11", ws4.cell(row=row_idx+offset, column=11).value)
        print("12", ws4.cell(row=row_idx+offset, column=12).value)
        print("13", ws4.cell(row=row_idx+offset, column=13).value)
        print("14", ws4.cell(row=row_idx+offset, column=14).value)
        print("15", ws4.cell(row=row_idx+offset, column=15).value)
        '''
        id = ws.cell(row=row_idx, column=3).value
        if id == None:
            continue
        name = ws.cell(row=row_idx, column=4).value
        if name == None:
            continue
        project_id = ws.cell(row=row_idx, column=6).value
        if project_id == None:
            continue
        str_date = ws.cell(row=row_idx, column=8).value
        if str_date == None or str_date == "____.__.__" or str_date == "________" or str_date == "":
            continue
        str_date = str_date.replace(".", "")
        date_begin = date(int(str_date[0:4]), int(str_date[4:6]), int(str_date[6:8]))
        str_date = ws.cell(row=row_idx, column=9).value
        if str_date == None or str_date == "____.__.__" or str_date == "________" or str_date == "":
            continue
        str_date = str_date.replace(".", "")
        date_end = date(int(str_date[0:4]), int(str_date[4:6]), int(str_date[6:8]))
        working_time = ws.cell(row=row_idx, column=11).value
        expenses_per_hour = ws.cell(row=row_idx, column=12).value
        expenses_holiday = ws.cell(row=row_idx, column=13).value
        expenses_monthly = ws.cell(row=row_idx, column=14).value
        expenses_extra = ws.cell(row=row_idx, column=15).value

        contract = Contract(TargetYear, id, name, project_id, date_begin, date_end, working_time, expenses_per_hour, expenses_monthly, expenses_holiday, expenses_extra)

        # Find the employee
        found = False
        for emp in Employees:
            if emp.id_ == id:
                emp.add(contract)
                print("Existing id and new contract", emp.id_, id, name)
                found = True
                break
        if not found:
            print("New id and new contract", id, name)

            employee = Employee(id, name)
            employee.add(contract)
            Employees.append(employee)

    print("Total employees", len(Employees))


    for employee in Employees:
        employee.show()

    wb.close()


def main():
    print("Hello, KetiManagerApp")

    func_init()
    #print("Current working directory:", os.getcwd())
    #print("Python path:", sys.path) 

    func_project_budget()

    func_employee()

main()
